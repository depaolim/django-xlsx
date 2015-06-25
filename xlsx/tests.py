from unittest import skipIf

from django.db import connection, models
from django.db.utils import IntegrityError
from django.test import TestCase, TransactionTestCase

import openpyxl as pyx

from .models import Iterator, load, _get


class Mock:
    def __init__(self, **kwargs):
        self.__dict__.update(kwargs)


class TestIterate(TestCase):

    def test_without_processor(self):
        rows = [
            [Mock(column="A", value="NAME"), Mock(column="B", value="ID_MAS")],
            [Mock(column="A", value="N1"), Mock(column="B", value="99")],
            [Mock(column="A", value="N2"), Mock(column="B", value="78")],
        ]
        REMAP_COL_TO_FIELD = {'NAME': "FIELD_A", "ID_MAS": "FIELD_B"}
        it = Iterator(REMAP_COL_TO_FIELD.get, lambda dr: None)
        i = it(rows, _get=lambda f, v: (f, v))
        self.assertEqual(i.next(), {"FIELD_A": "N1", "FIELD_B": "99"})
        self.assertEqual(i.next(), {"FIELD_A": "N2", "FIELD_B": "78"})
        self.assertRaises(StopIteration, i.next)

    def test_with_processor(self):
        def dummy_process(dr):
            if dr["FIELD_A"] == "N1":
                dr["FIELD_B"] = "N1_{}".format(dr["FIELD_B"])

        rows = [
            [Mock(column="A", value="NAME"), Mock(column="B", value="ID_MAS")],
            [Mock(column="A", value="N1"), Mock(column="B", value="99")],
            [Mock(column="A", value="N2"), Mock(column="B", value="78")],
        ]
        REMAP_COL_TO_FIELD = {'NAME': "FIELD_A", "ID_MAS": "FIELD_B"}
        it = Iterator(REMAP_COL_TO_FIELD.get, dummy_process)
        i = it(rows, _get=lambda f, v: (f, v))
        self.assertEqual(i.next(), {"FIELD_A": "N1", "FIELD_B": "N1_99"})
        self.assertEqual(i.next(), {"FIELD_A": "N2", "FIELD_B": "78"})
        self.assertRaises(StopIteration, i.next)


class SampleMaster(models.Model):
    pass


class SampleMasterNatural(models.Model):
    key = models.CharField(max_length=10, unique=True)


class SampleDetail(models.Model):
    master = models.ForeignKey(SampleMaster, null=True)
    master_natural = models.ForeignKey(SampleMasterNatural, null=True)
    name = models.CharField(max_length=50)

    XLSX_2_FIELDS = {
        "ID_MAS": master,
        "MAS_N": (master_natural, "key"),
        "NAME": name,
    }


def sample_fixture(test_case):
    test_case.m1 = SampleMaster()
    test_case.m1.save()
    test_case.m2 = SampleMaster()
    test_case.m2.save()
    SampleDetail(master=test_case.m1, name='foo').save()
    SampleDetail(master=test_case.m2, name='bar').save()


class TestFieldPair(TestCase):

    def setUp(self):
        self.smn1 = SampleMasterNatural(key='K1')
        self.smn1.save()

    def test_field_pair(self):
        field_specs = SampleDetail.XLSX_2_FIELDS["MAS_N"]
        self.assertEquals(
            _get(field_specs, "K1"), ("master_natural_id", self.smn1.pk))


class TestSampleModelUnit(TestCase):

    setUp = sample_fixture

    def test_load(self):
        rows = [
            [Mock(column="A", value="NAME"), Mock(column="B", value="ID_MAS")],
            [Mock(column="A", value="N1"), Mock(column="B", value=self.m1.pk)],
            [Mock(column="A", value="N2"), Mock(column="B", value=self.m2.pk)],
        ]

        load(SampleDetail, rows)
        cols = ('master', 'name')
        self.assertItemsEqual(
            SampleDetail.objects.order_by(*cols).values_list(*cols), [
                (self.m1.pk, "N1"),
                (self.m2.pk, "N2"),
            ])

    def test_load_with_preprocess(self):
        def dummy_preprocess(dr):
            dr["name"] = "{}_{}".format(dr["name"], dr["master_id"])

        rows = [
            [Mock(column="A", value="NAME"), Mock(column="B", value="ID_MAS")],
            [Mock(column="A", value="N1"), Mock(column="B", value=self.m1.pk)],
            [Mock(column="A", value="N2"), Mock(column="B", value=self.m2.pk)],
        ]

        load(SampleDetail, rows, dummy_preprocess)
        cols = ('master', 'name')
        self.assertItemsEqual(
            SampleDetail.objects.order_by(*cols).values_list(*cols), [
                (self.m1.pk, "N1_{}".format(self.m1.pk)),
                (self.m2.pk, "N2_{}".format(self.m2.pk)),
            ])

    def test_load_with_natural(self):
        SampleMasterNatural(key="MN1").save()
        SampleMasterNatural(key="MN2").save()

        rows = [
            [Mock(column="A", value="NAME"), Mock(column="B", value="MAS_N")],
            [Mock(column="A", value="N1"), Mock(column="B", value="MN1")],
            [Mock(column="A", value="N2"), Mock(column="B", value="MN2")],
        ]

        load(SampleDetail, rows)
        cols = ('master_natural__key', 'name')
        self.assertItemsEqual(
            SampleDetail.objects.order_by(*cols).values_list(*cols), [
                ("MN1", "N1"),
                ("MN2", "N2"),
            ])


class TestSensorVehicleForeignKeys(TransactionTestCase):

    @skipIf(
        connection.vendor == 'sqlite',
        "This test not valid for SQLite")
    def test_load_non_existent_vehicle(self):
        rows = [
            [Mock(column="A", value="ID_MAS")],
            [Mock(column="A", value="999")]
        ]
        with self.assertRaisesRegexp(
                IntegrityError, "master_id.=.999. is not present in table"):
            load(SampleDetail, rows)


class TestSampleModelIntegration(TestCase):

    SAMPLE_XLSX = "sample.xlsx"
    SHEET_NAME = "DATA"

    def setUp(self):
        sample_fixture(self)

        # create xlsx file
        wb = pyx.Workbook()
        ws = wb.active
        ws.title = self.SHEET_NAME
        ws["A1"] = "NAME"
        ws["B1"] = "ID_MAS"
        ws["C1"] = "UNWANTED COLUMN"
        ws["A2"] = "one-name"
        ws["B2"] = unicode(self.m1.pk)
        ws["A3"] = "another-name"
        ws["B3"] = unicode(self.m1.pk)
        ws["C3"] = "unwanted value"
        wb.save(self.SAMPLE_XLSX)

    def test_sample(self):
        wb = pyx.load_workbook(self.SAMPLE_XLSX)
        ws = wb[self.SHEET_NAME]
        self.assertEquals(len(ws.rows), 3)
        load(SampleDetail, ws.rows)
        cols = ('master', 'name')
        self.assertItemsEqual(
            SampleDetail.objects.order_by(*cols).values_list(*cols), [
                (self.m1.pk, "one-name"),
                (self.m1.pk, "another-name"),
            ])
